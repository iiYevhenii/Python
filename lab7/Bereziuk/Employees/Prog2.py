import csv
import openpyxl
from datetime import datetime
import os

def calculate_age(birth_date_str):
    try:
        birth_date = datetime.strptime(birth_date_str, "%Y-%m-%d")
        today = datetime.today()
        age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        return age
    except ValueError:
        return 0

def get_age_category(age):
    if age < 18:
        return "younger_18"
    elif 18 <= age <= 45:
        return "18-45"
    elif 45 < age <= 70:
        return "45-70"
    else:
        return "older_70"

def main():
    csv_file = 'employees.csv'
    xlsx_file = 'employees.xlsx'

    if not os.path.exists(csv_file):
        print("Повідомлення про відсутність, або проблеми при відкритті файлу CSV")
        return

    try:
        wb = openpyxl.Workbook()
        
        sheets = {
            "all": wb.active,
            "younger_18": wb.create_sheet("younger_18"),
            "18-45": wb.create_sheet("18-45"),
            "45-70": wb.create_sheet("45-70"),
            "older_70": wb.create_sheet("older_70")
        }
        sheets["all"].title = "all"

        cat_header = ["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"]
        for name, ws in sheets.items():
            if name != "all":
                ws.append(cat_header)

        counters = {"younger_18": 1, "18-45": 1, "45-70": 1, "older_70": 1}

        with open(csv_file, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            sheets["all"].append(reader.fieldnames)

            for row in reader:
                sheets["all"].append([row[h] for h in reader.fieldnames])

                age = calculate_age(row["Дата народження"])
                category = get_age_category(age)

                sheets[category].append([
                    counters[category],
                    row["Прізвище"],
                    row["Ім’я"],
                    row["По батькові"],
                    row["Дата народження"],
                    age
                ])
                counters[category] += 1

        wb.save(xlsx_file)
        print("Ok")

    except Exception:
        print("Повідомлення про неможливість створення XLSX файлу")

if __name__ == "__main__":
    main()